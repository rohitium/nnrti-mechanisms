#!/usr/bin/env python3
"""Build Supplementary Table 3 from the two binding-energetics protocols.

Two protocols, deliberately named apart because they are not the same quantity:

- **ddE** -- the MM/GBSA endpoint score. An interaction *energy* from fixed
  trajectory snapshots, with no configurational entropy and an implicit-solvent
  polar term. Cheap, computed for every genotype.
- **ddG** -- the pmx non-equilibrium FEP binding free energy. An equilibrium
  free-energy difference from alchemical switching. Expensive, and the
  quantity that actually corresponds to binding.

Reporting both in one table keeps the comparison explicit rather than letting a
reader assume the MM/GBSA column is a free energy.
"""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SUMMARY_COMPONENTS = [
    ("ddg", "ddE total"),
    ("ddg_vdw", "ddE van der Waals"),
    ("ddg_electrostatic", "ddE electrostatic"),
    ("ddg_gb", "ddE GB polar solvation"),
    ("ddg_sa", "ddE nonpolar SA"),
]

#: pmx NEQ FEP columns, joined on genotype.
FEP_COLUMNS = [
    ("ddg_bind_kcal", "ddG bind (kcal/mol)"),
    ("sem_kcal", "ddG SEM (kcal/mol)"),
]

#: MMGBSA sheet. One row per replicate, each energy paired with its SEM over the
#: sampled snapshots WITHIN that replicate -- a different quantity from the
#: across-replicate SEM reported in Table 2. WT-reference columns are omitted
#: because they would repeat one constant on every row; WT's own three replicates
#: are listed instead, so ddE for any row is that row's energy minus the WT mean.
DETAIL_COLUMNS = [
    ("mutation", "Mutation"),
    ("replicate", "Replicate"),
    ("dor_fold_reduction", "DOR fold reduction"),
    ("mmgbsa_snapshots", "MM/GBSA snapshots"),
    ("binding_dg", "Total Energy (kcal/mol)"),
    ("binding_dg_sem", "Total Energy SEM (kcal/mol)"),
    ("binding_dg_vdw", "van der Waals (kcal/mol)"),
    ("binding_dg_vdw_sem", "van der Waals SEM (kcal/mol)"),
    ("binding_dg_electrostatic", "Electrostatic (kcal/mol)"),
    ("binding_dg_electrostatic_sem", "Electrostatic SEM (kcal/mol)"),
    ("binding_dg_gb", "GB polar solvation (kcal/mol)"),
    ("binding_dg_gb_sem", "GB polar solvation SEM (kcal/mol)"),
    ("binding_dg_sa", "Nonpolar SA (kcal/mol)"),
    ("binding_dg_sa_sem", "Nonpolar SA SEM (kcal/mol)"),
]


def repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def parse_args() -> argparse.Namespace:
    root = repo_root()
    parser = argparse.ArgumentParser(
        description="Generate manuscript/Supplementary-Table-3.xlsx from WT-referenced energetic shifts."
    )
    parser.add_argument(
        "--ddg-csv",
        type=Path,
        default=root / "results/analysis/binding_energy/tables/ddg_full.csv",
        help="Replicate-level WT-referenced MM/GBSA table.",
    )
    parser.add_argument(
        "--panel-csv",
        type=Path,
        default=root / "results/analysis/dor_susceptibility_bar_chart/tables/dor_susceptibility_values.csv",
        help="Current manuscript mutation panel and DOR fold reductions.",
    )
    parser.add_argument(
        "--fep-csv",
        type=Path,
        default=root / "results/analysis/fep_pmx/panel_discussion_tiers.csv",
        help="pmx NEQ FEP panel (ddG_bind per genotype, with SEM and reporting tier).",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=root / "manuscript/Supplementary-Table-3.xlsx",
        help="Output workbook path.",
    )
    parser.add_argument(
        "--expected-replicates",
        type=int,
        default=3,
        help="Expected number of production replicates per mutation.",
    )
    return parser.parse_args()


def sem(values: pd.Series) -> float:
    numeric = pd.to_numeric(values, errors="coerce").dropna().to_numpy(dtype=float)
    if numeric.size <= 1:
        return 0.0
    return float(np.std(numeric, ddof=1) / np.sqrt(numeric.size))


def load_panel(panel_csv: Path) -> pd.DataFrame:
    panel = pd.read_csv(panel_csv)
    required = {"mutation", "dor_fold_reduction"}
    missing = required.difference(panel.columns)
    if missing:
        raise ValueError(f"{panel_csv} is missing required columns: {sorted(missing)}")
    panel = panel[["mutation", "dor_fold_reduction"]].copy()
    panel["mutation"] = panel["mutation"].astype(str)
    return panel


def load_fep(fep_csv: Path) -> pd.DataFrame:
    """pmx FEP ddG per genotype, keyed for joining onto the ddE panel."""
    if not fep_csv.exists():
        return pd.DataFrame(columns=["mutation", *[c for c, _l in FEP_COLUMNS]])
    fep = pd.read_csv(fep_csv)
    if "genotype" not in fep.columns:
        raise ValueError(f"{fep_csv} has no 'genotype' column")
    fep = fep.rename(columns={"genotype": "mutation"})
    fep["mutation"] = fep["mutation"].astype(str).str.strip()
    # The tier file carries ddG/SEM/fold/tier but not replicate counts; take those
    # from the panel table alongside it.
    if "n_reps" not in fep.columns:
        panel_ddg = fep_csv.with_name("panel_ddg.csv")
        if panel_ddg.exists():
            counts = pd.read_csv(panel_ddg).rename(columns={"genotype": "mutation"})
            counts["mutation"] = counts["mutation"].astype(str).str.strip()
            fep = fep.merge(counts[["mutation", "n_reps"]], on="mutation", how="left")
        else:
            fep["n_reps"] = pd.NA
    keep = ["mutation", *[c for c, _l in FEP_COLUMNS if c in fep.columns]]
    if "dor_fold_reduction" in fep.columns:
        keep.insert(1, "dor_fold_reduction")
    return fep[keep].drop_duplicates(subset="mutation")


def load_details(ddg_csv: Path, panel: pd.DataFrame, expected_replicates: int) -> pd.DataFrame:
    ddg = pd.read_csv(ddg_csv)
    required = {"mutation", "replicate", "mmgbsa_snapshots", *[column for column, _label in SUMMARY_COMPONENTS]}
    required.update(column for column, _label in DETAIL_COLUMNS if column in ddg.columns)
    missing = required.difference(ddg.columns)
    if missing:
        raise ValueError(f"{ddg_csv} is missing required columns: {sorted(missing)}")

    panel_order = panel["mutation"].tolist()
    # WT is not a panel genotype but is carried so the reference spread is visible.
    filtered = ddg[ddg["mutation"].isin([*panel_order, "WT"])].copy()
    missing_mutations = sorted(set(panel_order).difference(filtered["mutation"].astype(str)))
    if missing_mutations:
        raise ValueError(f"No energetic rows found for panel mutations: {missing_mutations}")

    replicate_counts = filtered.groupby("mutation", observed=True)["replicate"].nunique()
    bad_counts = replicate_counts[replicate_counts != expected_replicates]
    if not bad_counts.empty:
        bad = ", ".join(f"{mutation}={count}" for mutation, count in bad_counts.items())
        raise ValueError(f"Unexpected replicate counts; expected {expected_replicates}: {bad}")

    fold_lookup = panel.set_index("mutation")["dor_fold_reduction"]
    filtered["dor_fold_reduction"] = filtered["mutation"].map(fold_lookup)
    # WT first, then the panel in its published order.
    filtered["mutation"] = pd.Categorical(
        filtered["mutation"], categories=["WT", *panel_order], ordered=True
    )
    filtered = filtered.sort_values(["mutation", "replicate"]).reset_index(drop=True)
    filtered["mutation"] = filtered["mutation"].astype(str)
    return filtered


def build_summary(details: pd.DataFrame) -> pd.DataFrame:
    grouped = details.groupby("mutation", sort=False)
    summary = grouped.agg(
        **{
            "DOR fold reduction": ("dor_fold_reduction", "first"),
            "n replicates": ("replicate", "nunique"),
        }
    ).reset_index()
    summary = summary.rename(columns={"mutation": "Mutation"})
    for column, label in SUMMARY_COMPONENTS:
        summary[f"{label} mean (kcal/mol)"] = grouped[column].mean().to_numpy(dtype=float)
        summary[f"{label} SEM (kcal/mol)"] = grouped[column].apply(sem).to_numpy(dtype=float)
    return summary


def attach_fep(summary: pd.DataFrame, fep: pd.DataFrame) -> pd.DataFrame:
    """Left-join the FEP ddG columns; genotypes without FEP stay blank."""
    if fep.empty:
        return summary
    merged = summary.merge(
        fep.rename(columns={"mutation": "Mutation", **{c: l for c, l in FEP_COLUMNS}}),
        on="Mutation",
        how="left",
    )
    return merged


def build_detail_table(details: pd.DataFrame) -> pd.DataFrame:
    return details[[column for column, _label in DETAIL_COLUMNS]].rename(columns=dict(DETAIL_COLUMNS))


def wt_reference_note(details: pd.DataFrame) -> str:
    """One-line description of the shared WT reference and its uncertainty."""
    labels = [
        ("wt_binding_dg", "wt_binding_dg_sem", "total"),
        ("wt_binding_dg_vdw", "wt_binding_dg_vdw_sem", "van der Waals"),
        ("wt_binding_dg_electrostatic", "wt_binding_dg_electrostatic_sem", "electrostatic"),
        ("wt_binding_dg_gb", "wt_binding_dg_gb_sem", "GB polar solvation"),
        ("wt_binding_dg_sa", "wt_binding_dg_sa_sem", "nonpolar SA"),
    ]
    parts = []
    for value_col, sem_col, label in labels:
        if value_col not in details.columns:
            continue
        value = float(pd.to_numeric(details[value_col], errors="coerce").dropna().iloc[0])
        if sem_col in details.columns:
            sem_value = float(pd.to_numeric(details[sem_col], errors="coerce").dropna().iloc[0])
            parts.append(f"{label} {value:.1f} +/- {sem_value:.1f}")
        else:
            parts.append(f"{label} {value:.1f}")
    if not parts:
        return ""
    return (
        "WT reference (mean +/- SEM over three WT production replicates, kcal/mol): "
        + "; ".join(parts)
        + ". These reference uncertainties are common to every row and are not included in the "
        "per-mutation SEM values above."
    )


def build_fep_sheet(fep: pd.DataFrame, panel: pd.DataFrame) -> pd.DataFrame:
    """FEP panel ordered to match the ddE sheet, with absent genotypes shown blank."""
    order = panel["mutation"].tolist()
    out = pd.DataFrame({"Mutation": order}).merge(
        fep.rename(columns={"mutation": "Mutation"}), on="Mutation", how="left"
    )
    extra = [m for m in fep["mutation"] if m not in set(order)]
    if extra:
        out = pd.concat(
            [out, fep[fep["mutation"].isin(extra)].rename(columns={"mutation": "Mutation"})],
            ignore_index=True,
        )
    # Genotypes with no completed FEP leg are left out entirely rather than shown
    # as a blank row; add them back once their ddG lands.
    out = out[out["ddg_bind_kcal"].notna()].reset_index(drop=True)
    rename = {c: l for c, l in FEP_COLUMNS}
    out = out.rename(columns=rename)
    cols = ["Mutation"] + [l for _c, l in FEP_COLUMNS if l in out.columns]
    if "dor_fold_reduction" in out.columns:
        out = out.rename(columns={"dor_fold_reduction": "DOR fold reduction"})
        cols.insert(1, "DOR fold reduction")
    return out[cols]


NOTES = [
    ("Units", "All energies are kcal/mol."),
    (
        "ddE (MMGBSA sheet)",
        "MM/GBSA endpoint score: an interaction ENERGY from fixed trajectory snapshots, with no "
        "configurational entropy and an implicit-solvent polar term. Not a free energy.",
    ),
    (
        "ddG (FEP sheet)",
        "pmx non-equilibrium FEP binding FREE ENERGY from alchemical switching. This is the quantity "
        "that corresponds to binding. ddE and ddG are different observables and should not be read as "
        "interchangeable.",
    ),
    (
        "ddE referencing",
        "Each mutant replicate is referenced to the mean of the three WT production replicates. WT's own "
        "three replicates are listed on the MMGBSA sheet; their ddE values are each replicate's deviation "
        "from that mean and therefore show the reference spread directly. WT-reference columns are omitted "
        "because they would repeat one constant on every row.",
    ),
    (
        "Two kinds of SEM",
        "On the MMGBSA sheet, SEM is over the sampled snapshots WITHIN one replicate. The across-replicate "
        "SEM -- the uncertainty quoted in the manuscript table -- is computed from the three replicate rows "
        "and is a different, larger quantity. On the FEP sheet, SEM is across the three pmx replicates.",
    ),
    (
        "Coverage",
        "The FEP sheet lists only genotypes with a completed pmx leg; any genotype still running is omitted "
        "rather than shown blank, and is added once its ddG lands.",
    ),
]


def style_workbook(output_xlsx: Path, mmgbsa_rows: int, fep_rows: int) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(output_xlsx)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(bottom=Side(style="thin", color="D1D5DB"))

    ws = wb["MMGBSA"]
    for cell in ws[1]:
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=2, max_row=mmgbsa_rows + 1):
        # WT is formatted identically to the mutants: it is another data row, not
        # a callout.
        for cell in row:
            cell.border = thin
            if cell.column > 4:
                cell.number_format = "0.0000"
        row[1].number_format = "0"
        row[2].number_format = "0.0"
        row[3].number_format = "0"
    ws.freeze_panes = "C2"
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16 if i == 1 else 13

    ws = wb["FEP"]
    for cell in ws[1]:
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    for row in ws.iter_rows(min_row=2, max_row=fep_rows + 1):
        for cell in row:
            cell.border = thin
            if isinstance(cell.value, float):
                cell.number_format = "0.000"
    ws.freeze_panes = "A2"
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18 if i == 1 else 15

    ws = wb["Notes"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows(min_row=1, max_row=len(NOTES)):
        row[0].font = Font(bold=True, color="1F2937")
        row[0].alignment = Alignment(vertical="top")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = 46

    wb.save(output_xlsx)


def write_workbook(mmgbsa: pd.DataFrame, fep_sheet: pd.DataFrame, output_xlsx: Path) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    notes = pd.DataFrame(NOTES, columns=["Field", "Description"])
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        mmgbsa.to_excel(writer, sheet_name="MMGBSA", index=False)
        fep_sheet.to_excel(writer, sheet_name="FEP", index=False)
        notes.to_excel(writer, sheet_name="Notes", index=False, header=False)
    style_workbook(output_xlsx, mmgbsa_rows=len(mmgbsa), fep_rows=len(fep_sheet))


def main() -> int:
    args = parse_args()
    panel = load_panel(args.panel_csv)
    details_source = load_details(args.ddg_csv, panel, args.expected_replicates)
    mmgbsa = build_detail_table(details_source)
    fep = load_fep(args.fep_csv)
    fep_sheet = build_fep_sheet(fep, panel)
    absent = sorted(set(panel["mutation"]).difference(fep_sheet["Mutation"]))
    if absent:
        print(f"note: omitted from FEP sheet (no completed leg): {', '.join(absent)}")
    write_workbook(mmgbsa, fep_sheet, args.output_xlsx)
    print(f"Wrote {args.output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
