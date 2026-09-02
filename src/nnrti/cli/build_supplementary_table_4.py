#!/usr/bin/env python3
"""Build Supplementary Table 4: per-replicate structural observables.

Why this exists
---------------
Every structural number quoted in the Discussion -- burial, interplanar angles,
named distances, pocket volume, per-moiety contacts -- was previously scattered
across several analysis CSVs that had been produced at different times over
different genotype subsets, and two of them computed the *same* quantity with
slightly different frame handling. A reader had no way to trace a number in the
text back to a value the pipeline actually produced.

This collects all of them into one workbook, one row per (genotype, replicate),
from a single pass over three canonical sources:

  results/analysis/mechanisms/mechanism_coordinates.csv        (per-frame -> per-replicate mean)
  results/analysis/mechanisms/dor_moiety_contacts_per_replicate.csv
  results/analysis/modern_md_suite/tables/pocket_volume_per_rep.csv

Replicate means are averaged, never pooled across frames, so the Summary sheet's
SEM is the replicate-to-replicate SEM quoted in the manuscript.

Usage
-----
    PYTHONPATH=. python -m nnrti.cli.build_supplementary_table_4
"""
from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

#: (source column, output label, decimals). Order defines the sheet layout.
FRAME_METRICS = [
    ("res103_bb_to_triazN", "Res103 C=O to DOR triazolinone N (A)", 2),
    ("res103_sc_to_dor_polar", "Res103 side-chain polar atoms to DOR (A)", 2),
    ("chl_ring_burial", "RT heavy-atom contacts < 4.0 A of chlorocyanophenyl ring", 1),
    ("y188_interplanar_deg", "Tyr188 / chlorocyanophenyl interplanar angle (deg)", 1),
    ("res179_to_dor_mindist", "Val179 to DOR minimum distance (A)", 2),
    ("dor_to_S105_mindist", "Ser105 to DOR minimum distance (A)", 2),
    ("dor_to_res106_mindist", "Residue 106 to DOR minimum distance (A)", 2),
]
MOIETY_METRIC = ("pyridinone", "Pyridinone moiety contacts < 4.0 A", 1)
VOLUME_METRIC = ("pocket_volume_mean", "NNIBP volume (A^3)", 1)

ALL_METRICS = FRAME_METRICS + [MOIETY_METRIC, VOLUME_METRIC]

#: Main-text Table 3 carries the same summary numbers under compact symbolic
#: headers, so the two can never drift apart. Keys are the Summary-sheet labels.
TABLE3_HEADERS = {
    "Res103 C=O to DOR triazolinone N (A)": "d(103 C=O\u00b7\u00b7\u00b7N4x)",
    "Res103 side-chain polar atoms to DOR (A)": "d(103 sc\u00b7\u00b7\u00b7DOR)",
    "RT heavy-atom contacts < 4.0 A of chlorocyanophenyl ring": "N(chl)",
    "Tyr188 / chlorocyanophenyl interplanar angle (deg)": "\u03b8(Y188/chl)",
    "Val179 to DOR minimum distance (A)": "d(V179\u00b7\u00b7\u00b7DOR)",
    "Ser105 to DOR minimum distance (A)": "d(S105\u00b7\u00b7\u00b7DOR)",
    "Residue 106 to DOR minimum distance (A)": "d(106\u00b7\u00b7\u00b7DOR)",
    "Pyridinone moiety contacts < 4.0 A": "N(pyr)",
    "NNIBP volume (A^3)": "V(NNIBP)",
}

DEFINITIONS = [
    ("Scope",
     "One row per genotype and production replicate. 19 genotypes plus WT, three "
     "independent 100 ns replicates each. F227C is an intermediate alchemical leg, "
     "simulated but not part of the reported panel; it is included for completeness."),
    ("Aggregation",
     "Per-replicate values are means over every analysed frame of that replicate. The "
     "Summary sheet averages the three replicate means and reports the standard error "
     "across replicates -- the same replicate-level SEM quoted in the manuscript. Frame "
     "counts differ between replicates; no frame was excluded or weighted."),
    ("Periodic boundaries",
     "Protein chains were made whole by MDTraj bond-graph traversal before any distance "
     "or contact was computed."),
    ("Contacts",
     "Counted as protein-ligand heavy-atom PAIRS within 4.0 A, so one RT atom close to "
     "several ligand atoms contributes once per pair."),
    ("Ring burial",
     "Computed over the atoms of the chlorocyanophenyl ring only, excluding its chloro "
     "and cyano substituents. The pyridinone moiety count, by contrast, is a moiety "
     "count: the ring plus its own exocyclic substituents. The two are therefore not "
     "on a common scale and should not be compared to each other."),
    ("Interplanar angle",
     "Ring planes from singular value decomposition of the ring atom coordinates; the "
     "angle is between the two plane normals. Undefined for Y188L, which has no Tyr188 "
     "ring; those cells are blank."),
    ("Named distances",
     "Residue-to-DOR entries are minimum heavy-atom distances between the residue and "
     "any DOR heavy atom. The Res103 C=O entry is a specific atom pair -- the residue "
     "103 main-chain carbonyl oxygen and the DOR triazolinone nitrogen -- not a minimum "
     "over the residue."),
    ("NNIBP volume",
     "Per frame on a 0.75 A cubic grid over a 10 A sphere centred on the Ca centroid of "
     "the sixteen pocket-lining residues (p66 L100, K101, K103, V106, T107, V108, V179, "
     "Y181, Y188, V189, G190, F227, W229, L234, Y318; p51 E138), counting grid points "
     "further from every protein heavy atom than that atom's van der Waals radius plus a "
     "1.4 A probe. Ligand-independent, so holo and apo are directly comparable."),
    ("Provenance",
     "Generated by src/analysis/cli/build_supplementary_table_4.py from "
     "results/analysis/mechanisms/mechanism_coordinates.csv, "
     "results/analysis/mechanisms/dor_moiety_contacts_per_replicate.csv and "
     "results/analysis/modern_md_suite/tables/pocket_volume_per_rep.csv."),
]

PANEL_ORDER = [
    "WT",
    "V106I", "K103N", "Y181C", "G190A",
    "Y318F", "V106A", "A98G+F227C", "V106I+F227C", "V106A+F227L",
    "Y188L", "V106A+P225H", "V106A+L234I", "K103N+M230L",
    "V106M", "G190S", "L100I+K103N", "K103N+P225H", "G190E",
    "F227C",
]

CATEGORY = {
    "WT": "Wild type",
    "V106I": "Susceptible", "K103N": "Susceptible", "Y181C": "Susceptible", "G190A": "Susceptible",
    "Y318F": "Resistant", "V106A": "Resistant", "A98G+F227C": "Resistant",
    "V106I+F227C": "Resistant", "V106A+F227L": "Resistant", "Y188L": "Resistant",
    "V106A+P225H": "Resistant", "V106A+L234I": "Resistant", "K103N+M230L": "Resistant",
    "V106M": "Uncertain", "G190S": "Uncertain", "L100I+K103N": "Uncertain",
    "K103N+P225H": "Uncertain", "G190E": "Uncertain",
    "F227C": "Alchemical intermediate",
}


def repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _order(frame: pd.DataFrame, col: str = "Genotype") -> pd.DataFrame:
    rank = {m: i for i, m in enumerate(PANEL_ORDER)}
    out = frame.copy()
    out["_r"] = out[col].map(lambda m: rank.get(m, len(rank)))
    return out.sort_values(["_r"] + ([ "Replicate"] if "Replicate" in out else [])).drop(columns="_r")


def build(root: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    coords = pd.read_csv(root / "results/analysis/mechanisms/mechanism_coordinates.csv")
    moiety = pd.read_csv(root / "results/analysis/mechanisms/dor_moiety_contacts_per_replicate.csv")
    volume = pd.read_csv(root / "results/analysis/modern_md_suite/tables/pocket_volume_per_rep.csv")

    frame_cols = [c for c, _l, _d in FRAME_METRICS]
    per_rep = (
        coords.groupby(["mutation", "replicate"])[frame_cols].mean().reset_index()
    )
    n_frames = coords.groupby(["mutation", "replicate"]).size().rename("n_frames").reset_index()
    per_rep = per_rep.merge(n_frames, on=["mutation", "replicate"], how="left")

    per_rep = per_rep.merge(
        moiety[["mutation", "replicate", MOIETY_METRIC[0]]], on=["mutation", "replicate"], how="left"
    )
    per_rep = per_rep.merge(
        volume[["mutation", "replicate", VOLUME_METRIC[0]]], on=["mutation", "replicate"], how="left"
    )

    detail = per_rep.rename(columns={"mutation": "Genotype", "replicate": "Replicate",
                                     "n_frames": "Frames analysed"})
    detail["Category"] = detail["Genotype"].map(CATEGORY)
    for src, label, dec in ALL_METRICS:
        detail[label] = detail[src].round(dec)
    detail = detail[["Category", "Genotype", "Replicate", "Frames analysed"]
                    + [l for _s, l, _d in ALL_METRICS]]
    detail = _order(detail)

    # Aggregate from the UNROUNDED per-replicate values, then round for display --
    # summarising the rounded column shifts the SEM (e.g. 0.074 -> 0.08).
    raw = per_rep.rename(columns={"mutation": "Genotype", "replicate": "Replicate"})
    raw = _order(raw)
    rows = []
    for gt, grp in raw.groupby("Genotype", sort=False):
        rec = {"Category": CATEGORY.get(gt, ""), "Genotype": gt,
               "Replicates": int(grp["Replicate"].nunique())}
        for src, label, dec in ALL_METRICS:
            vals = pd.to_numeric(grp[src], errors="coerce").dropna()
            if vals.empty:
                rec[label] = ""
                continue
            m = vals.mean()
            s = vals.std(ddof=1) / np.sqrt(len(vals)) if len(vals) > 1 else 0.0
            # A SEM that rounds to zero is real (replicates agreeing to better than
            # the displayed precision) but "± 0.00" reads as a formatting fault.
            # A SEM that rounds to zero at the column's precision is real
            # (replicates agreeing more closely than the displayed digits), but
            # "± 0.00" reads as a formatting fault -- carry one extra digit.
            edec = dec
            while edec < dec + 3 and round(s, edec) == 0.0:
                edec += 1
            rec[label] = f"{m:.{dec}f} ± {s:.{edec}f}"
        rows.append(rec)
    summary = _order(pd.DataFrame(rows))
    return detail, summary


def style(path: Path, sheets: dict[str, int]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    head_fill = PatternFill("solid", fgColor="DDE5F0")
    thin = Side(style="thin", color="BBBBBB")
    for name, ncols in sheets.items():
        ws = wb[name]
        ws.freeze_panes = "D2" if name == "Per-replicate" else "C2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = head_fill
            cell.alignment = Alignment(wrap_text=True, vertical="bottom")
            cell.border = Border(bottom=thin)
        ws.row_dimensions[1].height = 58
        for i in range(1, ncols + 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 14 if i > 3 else 22
    ws = wb["Definitions"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 108
    for row in ws.iter_rows():
        row[0].font = Font(bold=True)
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def main() -> int:
    root = repo_root()
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--output-xlsx", type=Path,
                    default=root / "manuscript/Supplementary-Table-4.xlsx")
    ap.add_argument("--table3-csv", type=Path,
                    default=root / "manuscript/Table-3-structural.csv")
    args = ap.parse_args()

    detail, summary = build(root)
    defs = pd.DataFrame(DEFINITIONS, columns=["Item", "Description"])

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.output_xlsx, engine="openpyxl") as w:
        summary.to_excel(w, sheet_name="Summary", index=False)
        detail.to_excel(w, sheet_name="Per-replicate", index=False)
        defs.to_excel(w, sheet_name="Definitions", index=False)
    style(args.output_xlsx, {"Summary": summary.shape[1], "Per-replicate": detail.shape[1]})
    print(f"Wrote {args.output_xlsx}  ({len(detail)} replicate rows, {len(summary)} genotypes)")

    # Main-text Table 3: the same summary, minus the Replicates column (constant
    # at 3), under compact headers.
    t3 = summary.drop(columns=["Replicates"]).rename(columns=TABLE3_HEADERS)
    t3 = t3[t3["Genotype"] != "F227C"]   # alchemical intermediate, not a panel genotype
    t3.to_csv(args.table3_csv, index=False)
    print(f"Wrote {args.table3_csv}  ({len(t3)} rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
